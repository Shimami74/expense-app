import pdfplumber
import re
import pandas as pd
import csv

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def load_dictionary():
    d = {}

    try:
        with open("name_dictionary.csv", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)

            for row in reader:
                d[row["alias"]] = row["real_name"]

    except:
        pass

    return d


# =========================
# PDF全文抽出
# =========================
def extract_text(pdf_path):

    text = ""

    with pdfplumber.open(pdf_path) as pdf:

        for page in pdf.pages:

            t = page.extract_text()

            if t:
                text += t + "\n"

    return text


# =========================
# 名前抽出
# =========================
def extract_name(text):

    m = re.search(r'■([^\s]+)', text)

    if m:
        return m.group(1).strip()

    return ""


# =========================
# 伝票番号
# =========================
def extract_voucher(text):

    m = re.search(r"G\d{9}", text)

    return m.group(0) if m else ""


# =========================
# 部門名
# =========================
def extract_department(line):

    m = re.search(r"\d{5}\s+(.+?)\s+諸口", line)

    if m:
        return m.group(1).strip()

    return ""


# =========================
# 行解析
# =========================
def parse_record(lines):

    """
    1レコード例

    2026/04/06
    G001949584
    1 71込 96061 技術　トンネル・基礎
    諸口 27,200 27,200
    71013 社外
    中央復建コンサルタンツ...■岡部正...
    """

    joined = " ".join(lines)

    # 計上日
    date = ""

    for l in lines:
        if re.match(r"\d{4}/\d{2}/\d{2}", l):
            date = l
            break

    voucher = extract_voucher(joined)

    # 行情報
    line_info = ""

    for l in lines:
        if re.match(r"^\d+\s+\S+\s+\d{5}", l):
            line_info = l
            break

    # 金額行
    amount_line = ""

    for l in lines:
        if "諸口" in l or "付替仮勘定" in l or "未払金" in l:
            amount_line = l
            break

    # 借方金額
    amount = None

    m = re.search(
        r"(?:諸口|付替仮勘定|未払金)\s+([\d,]+)\s+[\d,]+",
        amount_line
    )

    if m:
        amount = int(m.group(1).replace(",", ""))

    # 行番号・税・部門コード
    行 = ""
    税 = ""
    部門コード = ""

    m = re.match(r"^(\d+)\s+(\S+)\s+(\d{5})", line_info)

    if m:
        行 = m.group(1)
        税 = m.group(2)
        部門コード = m.group(3)

    部門名 = extract_department(line_info)

    name = extract_name(joined)

    return {
        "ページ": "",
        "計上行": date,
        "伝票番号": voucher,
        "相手補助科目コード": "",
        "相手補助科目": "",
        "行": 行,
        "税": 税,
        "部門コード": 部門コード,
        "部門名": 部門名,
        "摘要": joined,
        "氏名": name,
        "借方金額": amount,
        "残高": ""
    }


# =========================
# レコード分割
# =========================
def split_records(text):

    lines = text.split("\n")

    records = []
    current = []

    inside = False

    for line in lines:

        line = line.strip()

        if not line:
            continue

        # ヘッダ除外
        if "総勘定元帳" in line:
            continue

        # 新レコード開始
        if re.match(r"\d{4}/\d{2}/\d{2}", line):

            if current:
                records.append(current)

            current = [line]
            inside = True

            continue

        if inside:
            current.append(line)

    if current:
        records.append(current)

    return records


# =========================
# メイン処理
# =========================
def process_pdf(pdf_path, output_excel):

    dictionary = load_dictionary()

    text = extract_text(pdf_path)

    records = split_records(text)

    rows = []

    unknown_rows = []

    total_amount = 0

    for r in records:

        row = parse_record(r)

        if not row["借方金額"]:
            continue

        # 名前辞書変換
        row["氏名"] = dictionary.get(
            row["氏名"],
            row["氏名"]
        )

        rows.append(row)

        total_amount += row["借方金額"]

        if not row["氏名"]:
            unknown_rows.append(row)

    # =========================
    df = pd.DataFrame(rows)

    columns = [
        "ページ",
        "計上行",
        "伝票番号",
        "相手補助科目コード",
        "相手補助科目",
        "行",
        "税",
        "部門コード",
        "部門名",
        "摘要",
        "氏名",
        "借方金額",
        "残高"
    ]

    df = df[columns]

    df["氏名"] = df["氏名"].replace("", "未抽出")

    # =========================
    # 集計
    # =========================
    summary = (
        df.groupby("氏名", as_index=False)["借方金額"]
        .sum()
        .rename(columns={"借方金額": "合計借方金額"})
        .sort_values(by="合計借方金額", ascending=False)
    )

    # =========================
    # Excel出力
    # =========================
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:

        df.to_excel(
            writer,
            sheet_name="抽出データ",
            index=False
        )

        summary.to_excel(
            writer,
            sheet_name="氏名別集計",
            index=False
        )

        pd.DataFrame(unknown_rows).to_excel(
            writer,
            sheet_name="未抽出",
            index=False
        )

    # =========================
    # openpyxl編集
    # =========================
    wb = load_workbook(output_excel)

    for ws in wb.worksheets:

        # ヘッダ太字
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # フィルター
        ws.auto_filter.ref = ws.dimensions

        # 列幅自動
        for col in ws.columns:

            max_len = 0

            col_letter = get_column_letter(col[0].column)

            for cell in col:

                try:
                    max_len = max(
                        max_len,
                        len(str(cell.value))
                    )
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_excel)

    return {
        "total_rows": len(df),
        "unknown_count": len(unknown_rows),
        "total_amount": total_amount
    }